import cv2
import os, csv
from PyQt5 import QtWidgets, QtGui
import sys
from PyQt5.QtWidgets import QApplication,QGraphicsScene,QGraphicsView,QGraphicsRectItem
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from datetime import datetime
from PyQt5.QtGui import QPainter, QColor
from PyQt5.QtGui import QColor, QPen
from PyQt5.QtCore import Qt, QRectF
from PyQt5.QtWidgets import  QMainWindow

image_ext = ('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')

def setWidth(sheet):
    """
    Function to set width of columns of Excel.

    Inputs:
    - None

    Outputs:
    - None
    """

    # Dictionary for column width
    col_width = {"B":6,"C":33,"D":12.29,"E":16.43,"F":11,"G":12.29}

    # Iterating dictionary for setting column width
    for column_letter, column_width in col_width.items():
        sheet.column_dimensions[column_letter].width = column_width

def excelMergeCells(start_row, start_column, end_row, end_column, sheet, value, color):
    """
    Function to do following task
    - Merge cells and set content and color of merged cell and center align text

    Inputs:
    - start_row
    - start_column
    - end_row
    - end_column
    - sheet
    - value
    - color

    Outputs:
    - None
    """

    ## Merging cell
    merge_range = f"{sheet.cell(row=start_row, column=start_column).coordinate}:{sheet.cell(row=end_row, column=end_column).coordinate}"
    sheet.merge_cells(merge_range)

    # Setting content of merge cells and Center align text
    cell = sheet.cell(row=start_row, column=start_column)
    alignment = Alignment(horizontal="center", vertical="center")
    cell.value = value
    cell.alignment = alignment

    # Setting border and color details
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    border = Border(
        left=Side(border_style="medium", color="000000"),
        right=Side(border_style="medium", color="000000"),
        top=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000"),
    )

    # Setting border and color of cells
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        for cell in row:
            cell.fill = fill
            cell.border = border

def fileSizeCalculate(original_folder_path, compressed_folder_path, image_file):
    """
    Function for determining size of original and compressed, the difference between the original and compressed image sizes, as well as the percentage of compression.
    
    Inputs:
    - original_folder_path : Original folder path
    - compressed_folder_path : Compressed folder path
    - image_file : Image Name

    Outputs:
    - image_file : Image Name
    - original_image_size_mb : Original image file size in MB
    - compressed_image_size_mb : Compressed image file size in MB
    - diff_size_mb : difference between the original and compressed image sizes in MB
    - diff_percentage : percentage of compression
    """

    # Calculating size of Original Image and converting in MB
    original_image_path = os.path.join(original_folder_path, image_file)
    original_image_size = os.path.getsize(original_image_path)
    original_image_size_mb = round(original_image_size / (1024 * 1024), 2)

    # Calculating size of Compressed Image and converting in MB
    compressed_image_path = os.path.join(compressed_folder_path, image_file)
    compressed_image_size = os.path.getsize(compressed_image_path)
    compressed_image_size_mb = round(compressed_image_size / (1024 * 1024), 2)

    # Calculating the difference between the original and compressed image sizes
    diff_size_mb = round((original_image_size_mb - compressed_image_size_mb), 2)

    # Calculating percentage of compression
    diff_percentage = round((diff_size_mb * 100) / original_image_size_mb, 2)

    return [image_file,original_image_size_mb,compressed_image_size_mb,diff_size_mb,diff_percentage]

def getImageCount(folder):
    """
    Function to get count of Image file count in folder

    Inputs:
    - folder : Folder path

    Outputs:
    - count : Count of image files
    """
    dir_list = os.listdir(folder)
    j = 0
    for image_file in dir_list:
        if not image_file.lower().endswith(image_ext):
            pass
        elif not os.path.isfile(os.path.join(folder,image_file)):
            pass
        else:
            j += 1
    print(j)
    return j

app = QApplication(sys.argv)


class Filecompressr(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        self.setWindowTitle("Bulk File Compress Tool")
        self.setWindowIcon(QtGui.QIcon("icon.png"))

        self.folderLabel = QtWidgets.QLabel("Folder:")
        self.folderEdit = QtWidgets.QLineEdit()
        self.folderButton = QtWidgets.QPushButton("Browse...")
        self.compressLabel = QtWidgets.QLabel("Compress %")
        self.compressEdit = QtWidgets.QLineEdit()
        self.compressButton = QtWidgets.QPushButton("Compress")

        layout = QtWidgets.QGridLayout()
        layout.addWidget(self.folderLabel, 0, 0)
        layout.addWidget(self.folderEdit, 0, 1)
        layout.addWidget(self.folderButton, 0, 2)
        layout.addWidget(self.compressLabel, 3, 0)
        layout.addWidget(self.compressEdit, 3, 1)
        layout.addWidget(self.compressButton, 5, 1)
        self.setLayout(layout)

        self.folderButton.clicked.connect(self.folderSelection)
        self.compressButton.clicked.connect(self.compressFiles)

    def folderSelection(self):
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Folder")
        self.folderEdit.setText(folder)

    def compressFiles(self):
        folder = self.folderEdit.text()
        compress_percentage = self.compressEdit.text()
        
        if compress_percentage != "":
            compress_percentage = int(compress_percentage)
            if compress_percentage < 10:
                QtWidgets.QMessageBox.warning(self, "compressd", "Compression % can not be less than 10%")
            elif compress_percentage > 90:
                QtWidgets.QMessageBox.warning(self, "compressd", "Compression % can not be greater than 90%")
            else:
                if folder and folder !='':
                    image_count = getImageCount(folder)
                    if(image_count!=0):
                        self.close()

                        original_total_size = 0
                        compressed_total_size = 0
                        compression = compress_percentage
                        original_folder_path = folder
                        actual_table_start_cell = 12

                        # Splitting Original path
                        path_split = os.path.split(original_folder_path)

                        # Creating Compressed images folder path
                        now = datetime.now()
                        dt_string = now.strftime("%d%m%Y%H%M%S")
                        compressed_folder_path = os.path.join(path_split[0], path_split[1] + "_" + dt_string + "_Compressed")
                        os.mkdir(compressed_folder_path)

                        # Creating excel file for storing compression details
                        
                        filename = "compression_details.xlsx"
                        excel_fle = os.path.join(compressed_folder_path, filename)
                        workbook = openpyxl.Workbook()
                        sheet = workbook.active

                        # Setting width of columns
                        setWidth(sheet)
                        
                        # Adding Source path in excel sheet
                        excelMergeCells(3,2,4, 7,sheet,"Source Path : " + original_folder_path,"76933C")

                        # Adding Destination path in excel sheet
                        excelMergeCells(6,2,7,7,sheet,"Destination Path : " + compressed_folder_path,"F79646")

                        # Getting today date
                        now = datetime.now()
                        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
                        excelMergeCells(9,2,9,7,sheet,"Date : " + dt_string,"2596be")

                        header = ["#","Image Name","Original Size","Compressed Size","Difference","Difference %"]
                        i = 0
                        # Setting center alignment,Border and color
                        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        alignment = Alignment(horizontal="center", vertical="center")
                        border = Border(
                            left=Side(border_style="medium", color="000000"),
                            right=Side(border_style="medium", color="000000"),
                            top=Side(border_style="medium", color="000000"),
                            bottom=Side(border_style="medium", color="000000"),
                        )

                        # Adding header to table 
                        for row in sheet.iter_rows(min_row=actual_table_start_cell, max_row=actual_table_start_cell, min_col=2, max_col=7):
                            for cell in row:
                                cell.fill = fill
                                cell.border = border
                                cell.value = header[i]
                                cell.alignment = alignment
                                i += 1

                        # Getting all the files from directory
                        dir_list = os.listdir(original_folder_path)
                        actual_image_count = 0
                        for image_file in dir_list:
                            
                            # Checing item in folder is file or not
                            if not os.path.isfile(os.path.join(original_folder_path,image_file)):
                                continue
                            
                            # Checking file is image file or not
                            if not image_file.lower().endswith(image_ext):
                                continue
                            
                            actual_image_count += 1

                            # Creating Original Image path
                            original_image_path = os.path.join(original_folder_path, image_file)

                            # Creating Compressed Image path
                            compressed_image_path = os.path.join(compressed_folder_path, image_file)

                            # Reading Image with the help of Open CV
                            image = cv2.imread(original_image_path)
                            
                            cv2.imwrite(compressed_image_path,image,[cv2.IMWRITE_JPEG_QUALITY, compression])
                            data = fileSizeCalculate(original_folder_path, compressed_folder_path, image_file)

                            final_data = [actual_image_count,data[0],str(data[1]) + " MB",str(data[2]) + " MB",str(data[3]) + " MB",data[4]]
                            i = 0
                            alignment = Alignment(horizontal="center", vertical="center")
                            for row in sheet.iter_rows(min_row=actual_table_start_cell + actual_image_count, max_row=actual_table_start_cell + actual_image_count, min_col=2, max_col=7):
                                for cell in row:
                                    cell.border = border
                                    cell.value = final_data[i]
                                    if i != 1:
                                        cell.alignment = alignment
                                    i += 1
                            
                            original_total_size += data[1]
                            compressed_total_size += data[2]
                        
                        # Calculating total compression details
                        compressed_total_size = round(compressed_total_size, 2)
                        original_total_size = round(original_total_size, 2)
                        diff_size_mb = round((original_total_size - compressed_total_size), 2)
                        diff_percentage = round((diff_size_mb * 100) / original_total_size, 2)
                        
                        # Adding Footer row in table
                        footer = ["","Total",str(original_total_size) + " MB",str(compressed_total_size) + " MB",str(diff_size_mb) + " MB",diff_percentage]
                        index_count = 0
                        actual_image_count += 1
                        fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
                        alignment = Alignment(horizontal="center", vertical="center")
                        for row in sheet.iter_rows(min_row=actual_table_start_cell + actual_image_count, max_row=actual_table_start_cell + actual_image_count, min_col=2, max_col=7):
                            for cell in row:
                                cell.fill = fill
                                cell.border = border
                                cell.value = footer[index_count]
                                cell.alignment = alignment
                                index_count += 1
                        
                        # Saving Excel workbook
                        workbook.save(excel_fle)
                        
                        # Show a message box to confirm that the files have been compressd
                        QtWidgets.QMessageBox.information(self, "compressd", "Files have been compressd")
                        
                    else:
                        # No image available in folder
                        QtWidgets.QMessageBox.information(self, "compressd", "No Image found in folder")
                        
                else:
                    # When folder is not selected
                    QtWidgets.QMessageBox.warning(self, "compressd", "Please select folder")
        else:
            # if Compression % is not entered
            QtWidgets.QMessageBox.warning(self, "compressd", "Please enter Compression % ")
        self.close()



if __name__ == "__main__":
    # Create an instance of the Filecompressr widget and show it
    compressr = Filecompressr()
    compressr.show()
    sys.exit(app.exec_())
